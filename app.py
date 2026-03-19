from flask import Flask, request, jsonify, send_from_directory, session
from flask_cors import CORS
import requests as req
import os, sqlite3, threading, time, uuid, json, re
from datetime import datetime
try:
    import pytz
    IST = pytz.timezone("Asia/Kolkata")
    def now_ist(): return datetime.now(IST)
except:
    def now_ist(): return datetime.utcnow()

app = Flask(__name__, static_folder=".")
CORS(app)
app.secret_key = os.environ.get("SECRET_KEY", "adword_v3_secret_x9k2m")

# ── CONFIG ────────────────────────────────────────────────
API_KEY   = os.environ.get("SMM_API_KEY", "877f4a9fcf5d5770b86f97867beea5bc")
API_URL   = "https://honestsmm.com/api/v2"
SERVICES  = {
    "1554": "SV4 FB Live Stream Views | 90 Min",
    "1236": "FB Live Stream Views | 90 Min | Instant Start"
}
USERS = {
    "admin":  {"password": os.environ.get("ADMIN_PASS",  "Admin@123"),  "role": "admin"},
    "rozmin": {"password": os.environ.get("ROZMIN_PASS", "Secure@123"), "role": "user"},
}
DB_PATH        = os.environ.get("DB_PATH", "adword.db")
CLAUDE_API_KEY = os.environ.get("CLAUDE_API_KEY", "")

# ── DATABASE ──────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS balance_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            txn_id TEXT UNIQUE NOT NULL,
            amount_usd REAL NOT NULL,
            added_by TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            order_num INTEGER NOT NULL,
            link TEXT NOT NULL,
            service_id TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            status TEXT NOT NULL,
            smm_order_id TEXT,
            smm_status TEXT DEFAULT 'pending',
            cost_usd REAL DEFAULT 0,
            created_at TEXT NOT NULL,
            username TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS job_sessions (
            id TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            link TEXT NOT NULL,
            service_id TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            total_orders INTEGER NOT NULL,
            gap_minutes INTEGER NOT NULL,
            placed INTEGER DEFAULT 0,
            success INTEGER DEFAULT 0,
            failed INTEGER DEFAULT 0,
            status TEXT DEFAULT 'running',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            show_name TEXT NOT NULL,
            page_url TEXT NOT NULL,
            service_id TEXT NOT NULL,
            days TEXT NOT NULL,
            start_time TEXT NOT NULL,
            total_views INTEGER NOT NULL,
            orders_count INTEGER NOT NULL,
            qty_per_order INTEGER NOT NULL,
            gap_minutes INTEGER NOT NULL,
            enabled INTEGER DEFAULT 1,
            created_by TEXT NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS scheduler_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            run_date TEXT NOT NULL,
            status TEXT NOT NULL,
            live_link TEXT,
            message TEXT,
            orders_placed INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );
        """)
        db.commit()

init_db()

# ── HELPERS ───────────────────────────────────────────────
def get_balance():
    with get_db() as db:
        row = db.execute("SELECT COALESCE(SUM(amount_usd),0) as bal FROM balance_logs").fetchone()
        return round(float(row["bal"]), 4)

def get_smm_balance():
    try:
        r = req.post(API_URL, data={"key": API_KEY, "action": "balance"}, timeout=10)
        d = r.json()
        if "balance" in d:
            return float(d["balance"])
    except: pass
    return None

# ── AI BRAIN — Claude verification ───────────────────────
def ai_verify_show(page_html, expected_show_name, live_link):
    """Ask Claude AI if the live show matches expected show name"""
    if not CLAUDE_API_KEY:
        return True, "AI verification skipped (no API key)"

    # Extract useful text from HTML
    title_match = re.search(r'<title>([^<]+)</title>', page_html)
    og_title    = re.search(r'og:title["\s]+content="([^"]+)"', page_html)
    og_desc     = re.search(r'og:description["\s]+content="([^"]+)"', page_html)

    page_title = title_match.group(1) if title_match else ""
    og_t       = og_title.group(1)   if og_title   else ""
    og_d       = og_desc.group(1)    if og_desc    else ""

    context = f"""
Page title: {page_title}
OG Title: {og_t}
OG Description: {og_d}
Live Link: {live_link}
Expected Show: {expected_show_name}
"""

    prompt = f"""You are verifying if a Facebook Live stream matches an expected TV show name.

Context:
{context}

Question: Is the currently live video likely the show "{expected_show_name}"?

Rules:
- Be lenient with spelling variations and language differences (show may be in French)
- If title/description contains similar words or the show name, say YES
- If it's clearly a completely different show, say NO
- If you cannot determine, say YES (give benefit of doubt)

Reply with ONLY: YES or NO, then a brief reason (1 line max)."""

    try:
        r = req.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": CLAUDE_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json"
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 100,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=15
        )
        d = r.json()
        reply = d["content"][0]["text"].strip()
        is_match = reply.upper().startswith("YES")
        return is_match, reply
    except Exception as e:
        return True, f"AI check failed ({e}) — proceeding anyway"

def scrape_with_html(page_url):
    """Scrape and return both link and raw HTML for AI verification"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        live_url = page_url.rstrip("/") + "/videos/live"
        r = req.get(live_url, headers=headers, timeout=15, allow_redirects=True)
        text = r.text

        link = None
        if "/videos/" in r.url and r.url != live_url:
            link = r.url
        if not link:
            og = re.search(r'property="og:url"\s+content="([^"]+/videos/\d+[^"]*)"', text)
            if og: link = og.group(1)
        if not link:
            vid = re.search(r'"video_id":"(\d+)"', text)
            if vid:
                page_name = page_url.rstrip("/").split("/")[-1]
                link = f"https://www.facebook.com/{page_name}/videos/{vid.group(1)}"
        if not link:
            perm = re.search(r'"permalink_url":"(https://www\.facebook\.com/[^/]+/videos/\d+[^"]*)"', text)
            if perm: link = perm.group(1).replace("\\/", "/")

        return link, text
    except:
        return None, ""


# ── SCRAPER (kept for backward compat) ───────────────────────────────────────────────
def scrape_live_link(page_url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        live_url = page_url.rstrip("/") + "/videos/live"
        r = req.get(live_url, headers=headers, timeout=15, allow_redirects=True)
        text = r.text

        if "/videos/" in r.url and r.url != live_url:
            return r.url

        og = re.search(r'property="og:url"\s+content="([^"]+/videos/\d+[^"]*)"', text)
        if og: return og.group(1)

        vid = re.search(r'"video_id":"(\d+)"', text)
        if vid:
            page_name = page_url.rstrip("/").split("/")[-1]
            return f"https://www.facebook.com/{page_name}/videos/{vid.group(1)}"

        perm = re.search(r'"permalink_url":"(https://www\.facebook\.com/[^/]+/videos/\d+[^"]*)"', text)
        if perm: return perm.group(1).replace("\\/", "/")

        return None
    except: return None

# ── BACKGROUND ORDER WORKER ───────────────────────────────
active_jobs = {}

def order_worker(session_id):
    stop_ev = active_jobs.get(session_id)
    with get_db() as db:
        job = db.execute("SELECT * FROM job_sessions WHERE id=?", (session_id,)).fetchone()
    if not job: return

    link=job["link"]; svc=job["service_id"]; qty=job["quantity"]
    total=job["total_orders"]; gap_sec=job["gap_minutes"]*60; username=job["username"]

    for i in range(1, total+1):
        if stop_ev and stop_ev.is_set(): break
        now = now_ist().strftime("%Y-%m-%d %H:%M:%S")
        try:
            r = req.post(API_URL, data={"key":API_KEY,"action":"add","service":svc,"link":link,"quantity":qty}, timeout=15)
            d = r.json()
            if "order" in d:
                with get_db() as db:
                    db.execute("""INSERT INTO orders (session_id,order_num,link,service_id,quantity,status,smm_order_id,created_at,username)
                        VALUES (?,?,?,?,?,?,?,?,?)""", (session_id,i,link,svc,qty,"success",str(d["order"]),now,username))
                    db.execute("UPDATE job_sessions SET placed=placed+1,success=success+1,updated_at=? WHERE id=?", (now,session_id))
                    db.commit()
            elif "error" in d:
                err=d["error"].lower()
                is_bal=any(w in err for w in ["balance","fund","credit","insufficient"])
                with get_db() as db:
                    db.execute("""INSERT INTO orders (session_id,order_num,link,service_id,quantity,status,smm_order_id,created_at,username)
                        VALUES (?,?,?,?,?,?,?,?,?)""", (session_id,i,link,svc,qty,"balance_low" if is_bal else "failed",d["error"],now,username))
                    db.execute("UPDATE job_sessions SET placed=placed+1,failed=failed+1,updated_at=? WHERE id=?", (now,session_id))
                    db.commit()
                if is_bal: break
        except Exception as e:
            now2=now_ist().strftime("%Y-%m-%d %H:%M:%S")
            with get_db() as db:
                db.execute("""INSERT INTO orders (session_id,order_num,link,service_id,quantity,status,smm_order_id,created_at,username)
                    VALUES (?,?,?,?,?,?,?,?,?)""", (session_id,i,link,svc,qty,"error",str(e),now2,username))
                db.execute("UPDATE job_sessions SET placed=placed+1,failed=failed+1,updated_at=? WHERE id=?", (now2,session_id))
                db.commit()
        if i<total and not (stop_ev and stop_ev.is_set()):
            stop_ev.wait(gap_sec)

    final=now_ist().strftime("%Y-%m-%d %H:%M:%S")
    final_status="stopped" if (stop_ev and stop_ev.is_set()) else "completed"
    with get_db() as db:
        db.execute("UPDATE job_sessions SET status=?,updated_at=? WHERE id=?", (final_status,final,session_id))
        db.commit()
    active_jobs.pop(session_id, None)

# ── SCHEDULER DAEMON ──────────────────────────────────────
scheduler_running = {}

def run_schedule(schedule_id):
    with get_db() as db:
        sch = db.execute("SELECT * FROM schedules WHERE id=?", (schedule_id,)).fetchone()
    if not sch: return

    now_t    = now_ist()
    run_date = now_t.strftime("%Y-%m-%d")
    log_time = now_t.strftime("%Y-%m-%d %H:%M:%S")
    show_name = sch["show_name"]

    # ── Step 1: Try to find live link (retry every 2 min for 15 min) ──
    live_link = None
    page_html = ""
    max_retries = 8   # 8 x 2min = 16 min window
    ai_verdict = ""

    for attempt in range(max_retries):
        if not scheduler_running.get(schedule_id): break

        found_link, html = scrape_with_html(sch["page_url"])

        if found_link:
            # ── Step 2: AI verification ──
            is_match, verdict = ai_verify_show(html, show_name, found_link)
            ai_verdict = verdict

            if is_match:
                live_link = found_link
                page_html = html
                break
            else:
                # Wrong show — log and keep retrying
                now_log = now_ist().strftime("%Y-%m-%d %H:%M:%S")
                with get_db() as db:
                    db.execute(
                        "INSERT INTO scheduler_logs (schedule_id,run_date,status,live_link,message,created_at) VALUES (?,?,?,?,?,?)",
                        (schedule_id, run_date, "wrong_show", found_link,
                         "AI: " + verdict + " | Retrying...", now_log))
                    db.commit()

        if attempt < max_retries - 1:
            # Wait 2 minutes before retry
            for _ in range(120):
                if not scheduler_running.get(schedule_id): break
                time.sleep(1)

    if not live_link:
        now_log = now_ist().strftime("%Y-%m-%d %H:%M:%S")
        with get_db() as db:
            db.execute(
                "INSERT INTO scheduler_logs (schedule_id,run_date,status,live_link,message,created_at) VALUES (?,?,?,?,?,?)",
                (schedule_id, run_date, "no_live", None,
                 "No matching live found after 16 min. AI: " + ai_verdict, now_log))
            db.commit()
        scheduler_running.pop(schedule_id, None)
        return

    log_time2 = now_ist().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        db.execute(
            "INSERT INTO scheduler_logs (schedule_id,run_date,status,live_link,message,orders_placed,created_at) VALUES (?,?,?,?,?,?,?)",
            (schedule_id, run_date, "running", live_link,
             "Live verified by AI: " + ai_verdict, 0, log_time2))
        log_id = db.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
        db.commit()

    total_placed=0
    gap_sec=sch["gap_minutes"]*60

    for i in range(sch["orders_count"]):
        if not scheduler_running.get(schedule_id): break
        try:
            r = req.post(API_URL, data={"key":API_KEY,"action":"add","service":sch["service_id"],"link":live_link,"quantity":sch["qty_per_order"]}, timeout=15)
            d = r.json()
            now2=now_ist().strftime("%Y-%m-%d %H:%M:%S")
            if "order" in d:
                total_placed+=1
                with get_db() as db:
                    db.execute("""INSERT INTO orders (session_id,order_num,link,service_id,quantity,status,smm_order_id,smm_status,created_at,username)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (f"sched_{schedule_id}_{run_date}",i+1,live_link,sch["service_id"],sch["qty_per_order"],"success",str(d["order"]),"pending",now2,sch["created_by"]))
                    db.execute("UPDATE scheduler_logs SET orders_placed=? WHERE id=?", (total_placed,log_id))
                    db.commit()
            elif "error" in d:
                if any(w in d["error"].lower() for w in ["balance","fund","credit","insufficient"]):
                    break
        except: pass
        if i < sch["orders_count"]-1 and scheduler_running.get(schedule_id):
            time.sleep(gap_sec)

    final2=now_ist().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        db.execute("UPDATE scheduler_logs SET status='done',message=?,orders_placed=? WHERE id=?",
                   (f"Done. {total_placed} orders placed.",total_placed,log_id))
        db.commit()
    scheduler_running.pop(schedule_id, None)

def scheduler_daemon():
    while True:
        try:
            now_t = now_ist()
            day_map={0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"}
            today=day_map[now_t.weekday()]
            cur_time=now_t.strftime("%H:%M")
            with get_db() as db:
                schedules=db.execute("SELECT * FROM schedules WHERE enabled=1").fetchall()
            for sch in schedules:
                days=json.loads(sch["days"])
                if today not in days: continue
                if sch["start_time"] != cur_time: continue
                if sch["id"] in scheduler_running: continue
                run_date=now_t.strftime("%Y-%m-%d")
                with get_db() as db:
                    already=db.execute("SELECT id FROM scheduler_logs WHERE schedule_id=? AND run_date=? AND status IN ('running','done')",(sch["id"],run_date)).fetchone()
                if already: continue
                scheduler_running[sch["id"]]=True
                threading.Thread(target=run_schedule,args=(sch["id"],),daemon=True).start()
        except: pass
        time.sleep(60)

threading.Thread(target=scheduler_daemon, daemon=True).start()

# ── AUTH DECORATORS ───────────────────────────────────────
def login_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**k):
        if not session.get("username"): return jsonify({"error":"Unauthorized"}),401
        return f(*a,**k)
    return d

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**k):
        if session.get("role")!="admin": return jsonify({"error":"Admin only"}),403
        return f(*a,**k)
    return d

# ── STATIC ROUTES ─────────────────────────────────────────
@app.route("/")
def index():
    if not session.get("username"): return send_from_directory(".","login.html")
    if session.get("role")=="admin": return send_from_directory(".","admin.html")
    return send_from_directory(".","dashboard.html")

@app.route("/scheduler.html")
def scheduler_html():
    if not session.get("username"): return jsonify({"error":"Unauthorized"}),401
    return send_from_directory("static","scheduler.html")

@app.route("/dashboard")
def dashboard():
    if not session.get("username"): return send_from_directory(".","login.html")
    return send_from_directory(".","dashboard.html")

# ── AUTH ──────────────────────────────────────────────────
@app.route("/login",methods=["POST"])
def login():
    data=request.json; u=data.get("username","").strip(); p=data.get("password","").strip()
    if u in USERS and USERS[u]["password"]==p:
        session["username"]=u; session["role"]=USERS[u]["role"]
        return jsonify({"success":True,"role":USERS[u]["role"]})
    return jsonify({"success":False,"error":"Invalid credentials"})

@app.route("/logout",methods=["POST"])
def logout():
    session.clear(); return jsonify({"success":True})

@app.route("/me")
@login_required
def me():
    return jsonify({"username":session["username"],"role":session["role"],"balance":get_balance(),"smm_balance":get_smm_balance()})

# ── ORDER ROUTES ──────────────────────────────────────────
@app.route("/start-session",methods=["POST"])
@login_required
def start_session():
    data=request.json; link=data.get("link","").strip(); svc=str(data.get("service_id","1554"))
    qty=data.get("quantity",0); total=data.get("total_orders",1); gap=data.get("gap_minutes",1)
    if "facebook.com" not in link: return jsonify({"error":"Valid Facebook link required"}),400
    if svc not in SERVICES: return jsonify({"error":"Invalid service"}),400
    if not (20<=int(qty)<=5000): return jsonify({"error":"Quantity must be 20-5000"}),400
    sid=str(uuid.uuid4()); now=now_ist().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        db.execute("""INSERT INTO job_sessions (id,username,link,service_id,quantity,total_orders,gap_minutes,placed,success,failed,status,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,0,0,0,'running',?,?)""",
            (sid,session["username"],link,svc,int(qty),int(total),int(gap),now,now))
        db.commit()
    stop_ev=threading.Event(); active_jobs[sid]=stop_ev
    threading.Thread(target=order_worker,args=(sid,),daemon=True).start()
    return jsonify({"success":True,"session_id":sid})

@app.route("/session-status/<sid>")
@login_required
def session_status(sid):
    with get_db() as db:
        job=db.execute("SELECT * FROM job_sessions WHERE id=?", (sid,)).fetchone()
        if not job: return jsonify({"error":"Not found"}),404
        orders=db.execute("SELECT order_num,status,smm_order_id,created_at FROM orders WHERE session_id=? ORDER BY order_num DESC LIMIT 10",(sid,)).fetchall()
    return jsonify({"session_id":job["id"],"status":job["status"],"placed":job["placed"],"success":job["success"],"failed":job["failed"],"total":job["total_orders"],"orders":[dict(o) for o in orders]})

@app.route("/stop-session/<sid>",methods=["POST"])
@login_required
def stop_session(sid):
    if sid in active_jobs: active_jobs[sid].set()
    return jsonify({"success":True})

@app.route("/my-sessions")
@login_required
def my_sessions():
    with get_db() as db:
        jobs=db.execute("SELECT * FROM job_sessions WHERE username=? ORDER BY created_at DESC LIMIT 20",(session["username"],)).fetchall()
    return jsonify([dict(j) for j in jobs])

# ── USER ROUTES ───────────────────────────────────────────
@app.route("/user/my-orders")
@login_required
def my_orders():
    with get_db() as db:
        rows=db.execute("SELECT * FROM orders WHERE username=? ORDER BY created_at DESC LIMIT 200",(session["username"],)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/user/stats")
@login_required
def user_stats():
    u=session["username"]
    with get_db() as db:
        to=db.execute("SELECT COUNT(*) as c FROM orders WHERE username=?",(u,)).fetchone()["c"]
        so=db.execute("SELECT COUNT(*) as c FROM orders WHERE username=? AND status='success'",(u,)).fetchone()["c"]
        tv=db.execute("SELECT COALESCE(SUM(quantity),0) as s FROM orders WHERE username=? AND status='success'",(u,)).fetchone()["s"]
        mo=db.execute("SELECT COUNT(*) as c FROM orders WHERE username=? AND strftime('%Y-%m',created_at)=strftime('%Y-%m','now')",(u,)).fetchone()["c"]
        mv=db.execute("SELECT COALESCE(SUM(quantity),0) as s FROM orders WHERE username=? AND strftime('%Y-%m',created_at)=strftime('%Y-%m','now') AND status='success'",(u,)).fetchone()["s"]
    return jsonify({"total_orders":to,"success_orders":so,"total_views":int(tv),"month_orders":mo,"month_views":int(mv)})

@app.route("/refresh-order-statuses",methods=["POST"])
@login_required
def refresh_order_statuses():
    u=session["username"]
    with get_db() as db:
        rows=db.execute("SELECT id,smm_order_id FROM orders WHERE username=? AND status='success' AND smm_order_id IS NOT NULL ORDER BY created_at DESC LIMIT 50",(u,)).fetchall()
    updated=0
    for row in rows:
        try:
            r=req.post(API_URL,data={"key":API_KEY,"action":"status","order":row["smm_order_id"]},timeout=8)
            d=r.json(); ss=d.get("status","").lower()
            if ss:
                with get_db() as db:
                    db.execute("UPDATE orders SET smm_status=? WHERE id=?",(ss,row["id"])); db.commit()
                updated+=1
        except: pass
    return jsonify({"success":True,"updated":updated})

# ── ADMIN ROUTES ──────────────────────────────────────────
@app.route("/admin/add-balance",methods=["POST"])
@login_required
@admin_required
def add_balance():
    data=request.json; txn_id=data.get("txn_id","").strip(); amount=float(data.get("amount",0))
    if not txn_id: return jsonify({"error":"Transaction ID required"}),400
    if amount<=0: return jsonify({"error":"Amount must be positive"}),400
    now=now_ist().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with get_db() as db:
            db.execute("INSERT INTO balance_logs (txn_id,amount_usd,added_by,created_at) VALUES (?,?,?,?)",(txn_id,amount,session["username"],now)); db.commit()
        return jsonify({"success":True,"new_balance":get_balance()})
    except sqlite3.IntegrityError:
        return jsonify({"error":"Transaction ID already used!"}),400

@app.route("/admin/balance-history")
@login_required
@admin_required
def balance_history():
    with get_db() as db:
        logs=db.execute("SELECT * FROM balance_logs ORDER BY created_at DESC").fetchall()
    return jsonify([dict(l) for l in logs])

@app.route("/admin/delete-balance/<int:log_id>",methods=["DELETE"])
@login_required
@admin_required
def delete_balance(log_id):
    with get_db() as db:
        db.execute("DELETE FROM balance_logs WHERE id=?",(log_id,)); db.commit()
    return jsonify({"success":True,"new_balance":get_balance()})

@app.route("/admin/all-orders")
@login_required
@admin_required
def all_orders():
    df=request.args.get("from",""); dt=request.args.get("to","")
    with get_db() as db:
        if df and dt:
            rows=db.execute("SELECT * FROM orders WHERE created_at>=? AND created_at<=? ORDER BY created_at DESC",(df+" 00:00:00",dt+" 23:59:59")).fetchall()
        else:
            rows=db.execute("SELECT * FROM orders ORDER BY created_at DESC LIMIT 500").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/admin/export-excel")
@login_required
@admin_required
def export_excel():
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except: return jsonify({"error":"openpyxl not installed"}),500
    df=request.args.get("from",""); dt=request.args.get("to","")
    with get_db() as db:
        if df and dt:
            orders=db.execute("SELECT * FROM orders WHERE created_at>=? AND created_at<=? ORDER BY created_at ASC",(df+" 00:00:00",dt+" 23:59:59")).fetchall()
        else:
            orders=db.execute("SELECT * FROM orders ORDER BY created_at ASC").fetchall()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Orders Report"
    hf=PatternFill("solid",fgColor="1a6ff5"); hfont=Font(color="FFFFFF",bold=True)
    headers=["#","Date","Username","Service","Link","Quantity","Order Status","SMM Status","SMM Order ID"]
    for ci,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=ci,value=h); cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal="center")
    for ri,row in enumerate(orders,2):
        ws.cell(row=ri,column=1,value=ri-1); ws.cell(row=ri,column=2,value=row["created_at"])
        ws.cell(row=ri,column=3,value=row["username"]); ws.cell(row=ri,column=4,value=f"#{row['service_id']} {SERVICES.get(row['service_id'],'')}") 
        ws.cell(row=ri,column=5,value=row["link"]); ws.cell(row=ri,column=6,value=row["quantity"])
        ws.cell(row=ri,column=7,value=row["status"]); ws.cell(row=ri,column=8,value=row.get("smm_status",""))
        ws.cell(row=ri,column=9,value=row["smm_order_id"])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or "")) for c in col)+4,50)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    from flask import send_file
    fname=f"orders_{df or 'all'}_{dt or 'all'}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/stats")
@login_required
@admin_required
def admin_stats():
    with get_db() as db:
        to=db.execute("SELECT COUNT(*) as c FROM orders").fetchone()["c"]
        so=db.execute("SELECT COUNT(*) as c FROM orders WHERE status='success'").fetchone()["c"]
        tv=db.execute("SELECT COALESCE(SUM(quantity),0) as s FROM orders WHERE status='success'").fetchone()["s"]
        mo=db.execute("SELECT COUNT(*) as c FROM orders WHERE strftime('%Y-%m',created_at)=strftime('%Y-%m','now')").fetchone()["c"]
        mv=db.execute("SELECT COALESCE(SUM(quantity),0) as s FROM orders WHERE strftime('%Y-%m',created_at)=strftime('%Y-%m','now') AND status='success'").fetchone()["s"]
        ba=db.execute("SELECT COALESCE(SUM(amount_usd),0) as s FROM balance_logs").fetchone()["s"]
    return jsonify({"total_orders":to,"success_orders":so,"total_views":int(tv),"this_month_orders":mo,"this_month_views":int(mv),"balance_added":round(float(ba),4),"current_balance":get_balance()})

# ── SCHEDULER ROUTES ──────────────────────────────────────
@app.route("/scheduler/list")
@login_required
def scheduler_list():
    with get_db() as db:
        rows=db.execute("SELECT * FROM schedules ORDER BY created_at DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/scheduler/add",methods=["POST"])
@login_required
def scheduler_add():
    data=request.json
    show_name=data.get("show_name","").strip()
    page_url=data.get("page_url","").strip()
    service_id=str(data.get("service_id","1554"))
    days=data.get("days",[])
    start_time=data.get("start_time","").strip()
    total_views=int(data.get("total_views",0))
    orders_count=int(data.get("orders_count",4))
    gap_minutes=int(data.get("gap_minutes",3))

    if not show_name: return jsonify({"error":"Show name required"}),400
    if not page_url: return jsonify({"error":"Page URL required"}),400
    if not days: return jsonify({"error":"Select at least one day"}),400
    if not start_time: return jsonify({"error":"Start time required"}),400
    if total_views<1: return jsonify({"error":"Views must be > 0"}),400

    qty_per_order=max(20,min(5000,total_views//orders_count))
    now=now_ist().strftime("%Y-%m-%d %H:%M:%S")
    with get_db() as db:
        db.execute("""INSERT INTO schedules (show_name,page_url,service_id,days,start_time,total_views,orders_count,qty_per_order,gap_minutes,enabled,created_by,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,1,?,?)""",
            (show_name,page_url,service_id,json.dumps(days),start_time,total_views,orders_count,qty_per_order,gap_minutes,session["username"],now))
        db.commit()
    return jsonify({"success":True})

@app.route("/scheduler/toggle/<int:sch_id>",methods=["POST"])
@login_required
def scheduler_toggle(sch_id):
    with get_db() as db:
        sch=db.execute("SELECT enabled FROM schedules WHERE id=?",(sch_id,)).fetchone()
        if not sch: return jsonify({"error":"Not found"}),404
        new_val=0 if sch["enabled"] else 1
        db.execute("UPDATE schedules SET enabled=? WHERE id=?",(new_val,sch_id)); db.commit()
    return jsonify({"success":True,"enabled":new_val})

@app.route("/scheduler/delete/<int:sch_id>",methods=["DELETE"])
@login_required
def scheduler_delete(sch_id):
    with get_db() as db:
        db.execute("DELETE FROM schedules WHERE id=?",(sch_id,)); db.commit()
    return jsonify({"success":True})

@app.route("/scheduler/logs")
@login_required
def scheduler_logs():
    with get_db() as db:
        logs=db.execute("""SELECT sl.*,s.show_name FROM scheduler_logs sl
            LEFT JOIN schedules s ON sl.schedule_id=s.id
            ORDER BY sl.created_at DESC LIMIT 50""").fetchall()
    return jsonify([dict(l) for l in logs])

@app.route("/scheduler/test-scrape",methods=["POST"])
@login_required
def test_scrape():
    data=request.json; page_url=data.get("page_url","").strip()
    if not page_url: return jsonify({"error":"URL required"}),400
    link=scrape_live_link(page_url)
    if link: return jsonify({"success":True,"live_link":link})
    return jsonify({"success":False,"message":"No live video found right now."})

@app.route("/scheduler/run-now/<int:sch_id>",methods=["POST"])
@login_required
def run_schedule_now(sch_id):
    if sch_id in scheduler_running:
        return jsonify({"error":"Already running"}),400
    scheduler_running[sch_id]=True
    threading.Thread(target=run_schedule,args=(sch_id,),daemon=True).start()
    return jsonify({"success":True,"message":"Schedule started manually!"})

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port)
